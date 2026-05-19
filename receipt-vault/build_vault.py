"""Receipt vault builder.

Pipeline:
  1) Gmail thread JSON dumps land in raw/<msg_id>.json (one message per file)
  2) extract.py walks raw/, parses each, inserts into receipts.db
  3) export.py renders receipts.xlsx + summary.txt from the DB

Run as a module to do everything from raw/ -> outputs:
  python3 build_vault.py
"""
from __future__ import annotations

import csv
import html
import json
import os
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

VAULT = Path("/home/tricia/receipt-vault")
RAW = VAULT / "raw"
DB_PATH = VAULT / "receipts.db"
XLSX_PATH = VAULT / "receipts.xlsx"
ERR_PATH = VAULT / "receipts_errors.csv"
SUMMARY_PATH = VAULT / "summary.txt"

VAULT.mkdir(parents=True, exist_ok=True)
RAW.mkdir(parents=True, exist_ok=True)

# ----- schema ----------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS receipts (
  gmail_msg_id TEXT PRIMARY KEY,
  date TEXT,
  time TEXT,
  merchant TEXT,
  merchant_address TEXT,
  category TEXT,
  items TEXT,
  subtotal REAL,
  tax REAL,
  total REAL,
  payment_method_last4 TEXT,
  source TEXT,
  gmail_link TEXT,
  notes TEXT,
  subject TEXT,
  sender TEXT
);
CREATE INDEX IF NOT EXISTS idx_date ON receipts(date);
CREATE INDEX IF NOT EXISTS idx_merchant ON receipts(merchant);
CREATE INDEX IF NOT EXISTS idx_category ON receipts(category);
"""

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    return conn

# ----- merchant -> category mapping -----------------------------------------

CATEGORY_RULES = [
    # (regex on sender or merchant text, category) -- evaluated top-to-bottom
    (r"\bmywallybot|sanford\s*(?:&|and)\s*sons|wallyguard|wally-?guard|found\.com", "business"),
    (r"\buspto|sc\.gov|scdor|\bsc\s+dor\b|\birs\b|treasury|\bdmv\b|dss\.sc\.gov|scdmh", "legal/gov"),
    (r"\bsantee\s+cooper|duke\s+energy|dominion\s+energy|sce&g|electric\s+coop", "utilities"),
    (r"at&t|\batt\.com|verizon|t-mobile|tmobile\.com|boost\s*mobile|boostmobile|spectrum|comcast|xfinity|cox\s+communications|frontier|centurylink", "utilities"),
    (r"hp\s*instant\s*ink|hpsmart", "tech/saas"),
    (r"hostinger|namecheap|godaddy|cloudflare|google\s+domains|google\s+workspace|squarespace|\bwix\b|shopify|canva", "tech/saas"),
    (r"anthropic|openai|cursor\.com|\bgithub\b|gitlab|vercel|netlify|render\.com|fly\.io|digitalocean|linode|\baws\b|amazon\s+web\s+services|bitdefender", "tech/saas"),
    (r"apple\.com|itunes|app\s+store|@email\.apple\.com", "tech/saas"),
    (r"google\s?play|google\s?pay|googleone|google\s+one", "tech/saas"),
    (r"netflix|hulu|disney\+|spotify|youtube\s+premium|paramount|peacock|\bhbo\b|max\.com", "tech/saas"),
    (r"\bmusc\b|cvs|walgreens|rite\s+aid|goodrx|good\s+rx|pharmacy|clinic|hospital|dental|optometr|vision|pediatric", "medical"),
    (r"usps|fedex|\bups\b|stamps\.com|pirate\s+ship", "shipping"),
    (r"geico|progressive|state\s+farm|allstate|usaa|farmers\s+insurance|liberty\s+mutual|nationwide", "household"),
    (r"home\s+depot|lowe'?s|ace\s+hardware|harbor\s+freight|menards", "household"),
    (r"\bamazon\b", "household"),
    (r"walmart", "food/groceries"),
    (r"target\.com|target\s+stores", "household"),
    (r"kroger|publix|food\s+lion|\baldi\b|whole\s+foods|trader\s+joe", "food/groceries"),
    (r"costco|sam'?s\s+club", "household"),
    (r"\bshell\b|\bexxon\b|\bchevron\b|\bbp\b|\bmobil\b|sunoco|circle\s+k|sheetz|wawa|murphy\s+usa|7-eleven|speedway", "gas"),
    (r"doordash|uber\s*eats|grubhub|chick-?fil-?a|mcdonald|burger\s+king|wendy|taco\s+bell|chipotle|panera|starbucks|dunkin|\bpizza\b|domino|papa\s+john", "food/groceries"),
    (r"school|daycare|kindercare|childcare", "kids"),
    (r"ebay|etsy|poshmark|mercari", "household"),
    (r"paypal|venmo|cash\s*app|zelle|\bstripe\b|\bsquare\b|\bzip\.co\b", "other"),
]

PRIMARY_MERCHANT_RULES = [
    # tighter merchant-name normalization -- evaluated top-to-bottom
    (r"\bmusc\b", "MUSC"),
    (r"hp\s*instant\s*ink|hpsmart", "HP Instant Ink"),
    (r"boost\s*mobile|boostmobile", "Boost Mobile"),
    (r"\bzip\.co\b|us-accounts\.zip", "Zip"),
    (r"\bfound\.com|noreply@found", "Found"),
    (r"2checkout|bitdefender", "Bitdefender (2checkout)"),
    (r"\bamazon\b", "Amazon"),
    (r"walmart", "Walmart"),
    (r"\btarget\b", "Target"),
    (r"\bebay\b", "eBay"),
    (r"\betsy\b", "Etsy"),
    (r"poshmark", "Poshmark"),
    (r"paypal", "PayPal"),
    (r"venmo", "Venmo"),
    (r"cash\s*app", "Cash App"),
    (r"apple\.com|@email\.apple\.com", "Apple"),
    (r"googleone|google\s+one|google\s+play|google\s+pay", "Google"),
    (r"\bstripe\.com\b|@stripe\.com", "Stripe"),
    (r"\bsquare\b", "Square"),
    (r"shopify", "Shopify"),
    (r"santee\s+cooper", "Santee Cooper"),
    (r"duke\s+energy", "Duke Energy"),
    (r"at&t|att\.com", "AT&T"),
    (r"verizon", "Verizon"),
    (r"t-mobile|tmobile\.com", "T-Mobile"),
    (r"\busps\b", "USPS"),
    (r"\bfedex\b", "FedEx"),
    (r"\bups\.com\b", "UPS"),
    (r"hostinger", "Hostinger"),
    (r"namecheap", "Namecheap"),
    (r"anthropic", "Anthropic"),
    (r"openai", "OpenAI"),
    (r"cloudflare", "Cloudflare"),
    (r"\buspto\b", "USPTO"),
    (r"scdor|sc\s+dor", "SC DOR"),
    (r"\birs\b", "IRS"),
    (r"canva", "Canva"),
    (r"spotify", "Spotify"),
    (r"netflix", "Netflix"),
    (r"hulu", "Hulu"),
    (r"\bcvs\b", "CVS"),
    (r"walgreens", "Walgreens"),
    (r"home\s+depot", "Home Depot"),
    (r"lowe'?s", "Lowe's"),
    (r"doordash", "DoorDash"),
    (r"uber\s*eats", "Uber Eats"),
    (r"grubhub", "Grubhub"),
    (r"starbucks", "Starbucks"),
]


def categorize(merchant: str, sender: str, subject: str) -> str:
    blob = " ".join([merchant or "", sender or "", subject or ""]).lower()
    for pat, cat in CATEGORY_RULES:
        if re.search(pat, blob):
            return cat
    return "other"


def normalize_merchant(sender: str, subject: str, fallback: str = "") -> str:
    blob = " ".join([sender or "", subject or "", fallback or ""]).lower()
    for pat, name in PRIMARY_MERCHANT_RULES:
        if re.search(pat, blob):
            return name
    # try to pull from sender display: "Foo Bar <foo@bar.com>" -> Foo Bar
    m = re.match(r"\s*([^<]+?)\s*<", sender or "")
    if m:
        return m.group(1).strip().title()
    if sender:
        # fallback to domain
        m = re.search(r"@([\w.-]+)", sender)
        if m:
            return m.group(1).split(".")[0].title()
    return fallback or "Unknown"


# ----- regex extractors ------------------------------------------------------

MONEY = r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})|[0-9]+\.[0-9]{2})"

PATTERNS_TOTAL = [
    re.compile(r"order\s+total[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"grand\s+total[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"total\s+(?:charged|paid|amount|due)[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"amount[^\$\d]{0,8}" + MONEY, re.I),
    re.compile(r"you\s+(?:paid|sent|were charged)[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"payment\s+amount[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"payment\s+of[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"received\s+your\s+" + MONEY + r"\s*(?:USD|payment)?", re.I),
    re.compile(r"processed\s+your\s+payment\s+of[^\$\d]{0,8}" + MONEY, re.I),
    re.compile(r"transfer\s+for[^\$\d]{0,8}" + MONEY, re.I),
    re.compile(r"charged\s+(?:you\s+)?" + MONEY, re.I),
    re.compile(r"your\s+" + MONEY + r"\s+(?:payment|order|charge)", re.I),
    re.compile(r"\btotal[^\$\d]{0,8}" + MONEY, re.I),
    re.compile(r"\$\s*([0-9]{1,3}(?:,[0-9]{3})*\.[0-9]{2})", re.I),  # fallback any $X.XX
]
PATTERNS_SUBTOTAL = [
    re.compile(r"subtotal[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"item\s+subtotal[^\$\d]{0,15}" + MONEY, re.I),
]
PATTERNS_TAX = [
    re.compile(r"(?:sales\s+)?tax[^\$\d]{0,15}" + MONEY, re.I),
    re.compile(r"estimated tax[^\$\d]{0,15}" + MONEY, re.I),
]
PATTERN_LAST4 = re.compile(r"(?:ending in|ending\s*[-:]?\s*|\*+|x{2,}|card\s*\*?)\s*(\d{4})\b", re.I)


def to_money(s: str) -> float | None:
    if s is None:
        return None
    s = s.replace(",", "").replace("$", "").strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def first_match(patterns, text):
    for pat in patterns:
        m = pat.search(text)
        if m:
            v = to_money(m.group(1))
            if v is not None and 0 < v < 1_000_000:
                return v
    return None


def strip_html(s: str) -> str:
    if not s:
        return ""
    # remove style/script
    s = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", s, flags=re.S | re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = html.unescape(s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


SUBJECT_TOTAL = re.compile(r"\$\s*" + MONEY)


def extract_receipt(msg: dict) -> dict | None:
    """Return a receipts row dict, or None to skip.

    msg shape: {date, id, sender, subject, snippet, plaintext_body, htmlBody, threadId}
    """
    msg_id = msg.get("id")
    if not msg_id:
        return None
    subject = msg.get("subject") or ""
    sender = msg.get("sender") or ""
    snippet = msg.get("snippet") or ""
    body_plain = msg.get("plaintext_body") or msg.get("plaintextBody") or ""
    body_html = msg.get("htmlBody") or msg.get("html_body") or ""
    body_text = body_plain or strip_html(body_html) or snippet
    text = "\n".join([subject, body_text])

    # quick reject: must look like a receipt
    has_money = bool(re.search(r"\$\s*\d", text) or re.search(r"\b\d+\.\d{2}\b", text))
    if not has_money:
        return None

    total = first_match(PATTERNS_TOTAL, text)
    subtotal = first_match(PATTERNS_SUBTOTAL, text)
    tax = first_match(PATTERNS_TAX, text)
    if total is None:
        # try subject (e.g. "Your $12.34 payment to Acme")
        m = SUBJECT_TOTAL.search(subject)
        if m:
            total = to_money(m.group(1))
    if total is None:
        return None  # no money found; bail

    last4 = None
    m = PATTERN_LAST4.search(text)
    if m:
        last4 = m.group(1)

    iso_date = msg.get("date") or ""
    try:
        dt = datetime.fromisoformat(iso_date.replace("Z", "+00:00"))
        date_s = dt.strftime("%Y-%m-%d")
        time_s = dt.strftime("%H:%M:%S")
    except Exception:
        date_s = iso_date[:10]
        time_s = iso_date[11:19]

    merchant = normalize_merchant(sender, subject)
    category = categorize(merchant, sender, subject)
    thread_id = msg.get("threadId") or msg_id
    link = f"https://mail.google.com/mail/u/0/#all/{thread_id}"

    # items: pull a short snippet from subject + first 200 chars
    items = (subject[:160]).strip()

    return {
        "gmail_msg_id": msg_id,
        "date": date_s,
        "time": time_s,
        "merchant": merchant,
        "merchant_address": "",
        "category": category,
        "items": items,
        "subtotal": subtotal,
        "tax": tax,
        "total": total,
        "payment_method_last4": last4,
        "source": "gmail",
        "gmail_link": link,
        "notes": "",
        "subject": subject,
        "sender": sender,
    }


# ----- ingest ---------------------------------------------------------------

COLUMNS = [
    "gmail_msg_id", "date", "time", "merchant", "merchant_address",
    "category", "items", "subtotal", "tax", "total",
    "payment_method_last4", "source", "gmail_link", "notes", "subject", "sender",
]


def upsert(conn, row: dict):
    placeholders = ",".join("?" for _ in COLUMNS)
    cols = ",".join(COLUMNS)
    update = ",".join(f"{c}=excluded.{c}" for c in COLUMNS if c != "gmail_msg_id")
    sql = (
        f"INSERT INTO receipts ({cols}) VALUES ({placeholders}) "
        f"ON CONFLICT(gmail_msg_id) DO UPDATE SET {update}"
    )
    conn.execute(sql, [row.get(c) for c in COLUMNS])


# reject patterns: not actual receipts even if they match the priority search
REJECT_SUBJECT_PATTERNS = [
    re.compile(r"payment\s+declined", re.I),
    re.compile(r"update\s+your\s+payment", re.I),
    re.compile(r"action\s+required", re.I),
    re.compile(r"payment\s+is\s+past\s+due", re.I),
    re.compile(r"we\s+can.?t\s+process\s+your\s+payment", re.I),
    re.compile(r"we\s+couldn.?t\s+process", re.I),
    re.compile(r"problem\s+with\s+your\s+payment", re.I),
    re.compile(r"failed", re.I),
    re.compile(r"final\s+reminder", re.I),
    re.compile(r"will\s+you\s+rate", re.I),
    re.compile(r"please\s+confirm\s+receipt", re.I),
    re.compile(r"auto[-\s]?reply|automatic\s+reply", re.I),
    re.compile(r"return\s+(initiated|request)", re.I),
    re.compile(r"refund", re.I),  # may revisit; usually paired with original receipt
]

REJECT_SENDER_PATTERNS = [
    re.compile(r"swiftloantrack|wealthora|airtalkwireless", re.I),
]


def is_rejected(subject: str, sender: str) -> str | None:
    for pat in REJECT_SUBJECT_PATTERNS:
        if pat.search(subject or ""):
            return f"subject_reject:{pat.pattern}"
    for pat in REJECT_SENDER_PATTERNS:
        if pat.search(sender or ""):
            return f"sender_reject:{pat.pattern}"
    return None


def iter_messages_from_file(path: Path):
    """Yield (msg_dict, source_tag) from a raw/*.json file.

    Supports two shapes:
      - search response: {"threads": [{"id":..., "messages":[{...}]}], ...}
      - single get_thread response: {"id":..., "messages":[{...}]}
      - single picked-out message: {"id":..., "subject":..., ...}
    """
    try:
        data = json.loads(path.read_text())
    except Exception as e:
        return
    if isinstance(data, dict) and "threads" in data:
        for t in data.get("threads") or []:
            for m in t.get("messages") or []:
                if "threadId" not in m:
                    m["threadId"] = t.get("id")
                yield m, "search"
        return
    if isinstance(data, dict) and "messages" in data:
        for m in data.get("messages") or []:
            if "threadId" not in m:
                m["threadId"] = data.get("id")
            yield m, "thread"
        return
    if isinstance(data, dict) and "id" in data:
        yield data, "msg"


def ingest_raw(verbose=True):
    """Walk raw/ and ingest. Dedups by gmail_msg_id (DB PK)."""
    conn = db()
    write_header = not ERR_PATH.exists()
    err_file = open(ERR_PATH, "a", newline="")
    err_writer = csv.writer(err_file)
    if write_header:
        err_writer.writerow(["msg_id", "subject", "sender", "reason"])

    n_in = 0
    n_out = 0
    n_err = 0
    n_reject = 0
    batch = 0
    seen = set()
    files = sorted(RAW.glob("*.json"))
    for f in files:
        for msg, tag in iter_messages_from_file(f):
            msg_id = msg.get("id")
            if not msg_id or msg_id in seen:
                continue
            seen.add(msg_id)
            n_in += 1
            subject = msg.get("subject", "") or ""
            sender = msg.get("sender", "") or ""
            reason = is_rejected(subject, sender)
            if reason:
                err_writer.writerow([msg_id, subject[:120], sender, reason])
                n_reject += 1
                continue
            try:
                row = extract_receipt(msg)
            except Exception as e:
                n_err += 1
                err_writer.writerow([msg_id, subject[:120], sender, f"extract:{e}"])
                continue
            if row is None:
                err_writer.writerow([msg_id, subject[:120], sender, "no_total"])
                continue
            upsert(conn, row)
            n_out += 1
            batch += 1
            if batch >= 50:
                conn.commit()
                batch = 0
    conn.commit()
    err_file.close()
    if verbose:
        print(f"ingest: scanned {n_in} msgs, wrote {n_out} receipts, rejected {n_reject}, errors {n_err}")
    return n_in, n_out, n_err


# ----- excel + summary -------------------------------------------------------

def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = db()
    cur = conn.execute(
        "SELECT date, time, merchant, merchant_address, category, items, "
        "subtotal, tax, total, payment_method_last4, source, gmail_msg_id, "
        "gmail_link, notes FROM receipts ORDER BY date DESC, time DESC"
    )
    rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts"
    headers = [
        "date", "time", "merchant", "merchant_address", "category", "items",
        "subtotal", "tax", "total", "payment_method_last4", "source",
        "gmail_msg_id", "gmail_link", "notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append(list(r))

    # dollar formatting on subtotal, tax, total (cols G, H, I)
    for col_letter in ("G", "H", "I"):
        for cell in ws[col_letter][1:]:
            cell.number_format = '"$"#,##0.00'

    # autofilter + widths
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    widths = [11, 9, 22, 26, 14, 50, 11, 9, 11, 8, 8, 18, 36, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # hyperlink the gmail_link column
    link_col = headers.index("gmail_link") + 1
    for row_idx in range(2, ws.max_row + 1):
        c = ws.cell(row=row_idx, column=link_col)
        if c.value:
            c.hyperlink = c.value
            c.font = Font(color="0563C1", underline="single")

    wb.save(XLSX_PATH)
    print(f"wrote {XLSX_PATH} ({len(rows)} rows)")
    return len(rows)


def write_summary():
    conn = db()
    rows = list(conn.execute(
        "SELECT date, merchant, category, total FROM receipts WHERE total IS NOT NULL"
    ))
    n = len(rows)
    total_spend = sum(r[3] or 0 for r in rows)
    by_cat = Counter()
    by_merch = Counter()
    for _date, merch, cat, tot in rows:
        by_cat[cat or "other"] += tot or 0
        by_merch[merch or "Unknown"] += tot or 0

    lines = []
    lines.append("# Receipt vault summary")
    lines.append(f"Generated: {datetime.utcnow().isoformat()}Z")
    lines.append(f"Total receipts: {n}")
    lines.append(f"Total $ spent:  ${total_spend:,.2f}")
    lines.append("")
    lines.append("## Spend by category")
    for cat, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        lines.append(f"  {cat:<18} ${amt:>12,.2f}")
    lines.append("")
    lines.append("## Top 10 merchants by spend")
    for merch, amt in by_merch.most_common(10):
        lines.append(f"  {merch:<28} ${amt:>12,.2f}")
    text = "\n".join(lines) + "\n"
    SUMMARY_PATH.write_text(text)
    print(text)
    return text


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "all"
    if cmd in ("ingest", "all"):
        ingest_raw()
    if cmd in ("export", "all"):
        export_xlsx()
        write_summary()
    if cmd == "init":
        db().close()
        print("db initialized")
